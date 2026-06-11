import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


TEXT_COLUMNS = (
    "text",
    "feedback",
    "comment",
    "comments",
    "content",
    "review",
    "review_text",
    "body",
    "description",
    "message",
    "conteudo",
    "texto",
)
ID_COLUMNS = ("id", "feedback_id", "source_id", "codigo", "reviewid", "review_id")
TARGET_COLUMNS = ("target", "technical_target", "alvo", "alvo_tecnico")
INTENT_COLUMNS = ("intent", "intention", "intencao")


@dataclass(frozen=True)
class CsvFeedback:
    source_id: str
    text: str
    target: str = ""
    intent: str = ""


@dataclass(frozen=True)
class CsvInspection:
    total_rows: int
    valid_rows: int
    empty_rows: int
    missing_text_rows: int
    fieldnames: tuple[str, ...]
    text_column: str
    id_column: str
    target_column: str
    intent_column: str
    delimiter: str
    file_format: str
    sheet_name: str
    warnings: tuple[str, ...]


@dataclass(frozen=True)
class _XlsxSchema:
    fieldnames: tuple[str, ...]
    text_column: str
    id_column: str
    target_column: str
    intent_column: str
    initial_rows: tuple[tuple[str, ...], ...]
    warnings: tuple[str, ...] = ()


def count_rows(path: Path, limit: int | None = None) -> int:
    if _is_excel_path(path):
        return inspect_csv(path, limit=limit).valid_rows

    count = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        reader = csv.DictReader(handle, dialect=_sniff_dialect(sample))
        for row in reader:
            if not any(row.values()):
                continue
            count += 1
            if limit and count >= limit:
                return limit
    return count


def inspect_csv(path: Path, limit: int | None = None) -> CsvInspection:
    if _is_excel_path(path):
        return _inspect_xlsx(path, limit=limit)

    total_rows = 0
    valid_rows = 0
    empty_rows = 0
    missing_text_rows = 0
    warnings = []

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        dialect = _sniff_dialect(sample)
        reader = csv.DictReader(handle, dialect=dialect)
        fieldnames = tuple(reader.fieldnames or ())
        normalized = {_normalize_column(name): name for name in fieldnames}
        text_column = _first_existing(normalized, TEXT_COLUMNS)
        id_column = _first_existing(normalized, ID_COLUMNS)
        target_column = _first_existing(normalized, TARGET_COLUMNS)
        intent_column = _first_existing(normalized, INTENT_COLUMNS)

        if text_column is None:
            available = ", ".join(fieldnames)
            raise ValueError(
                "CSV sem coluna de texto reconhecida. Use uma coluna chamada text, feedback, comment, "
                f"content, review, description ou message. Colunas encontradas: {available}"
            )

        if id_column is None:
            warnings.append("Nenhuma coluna de ID reconhecida; o numero da linha sera usado como identificador.")
        if target_column is None:
            warnings.append("Nenhuma coluna de alvo tecnico reconhecida; o alvo sera inferido pelo pipeline.")
        if intent_column is None:
            warnings.append("Nenhuma coluna de intencao reconhecida; a intencao sera inferida pelo pipeline.")

        for row in reader:
            if limit and valid_rows >= limit:
                break
            if not any(row.values()):
                empty_rows += 1
                continue
            total_rows += 1
            text = (row.get(text_column) or "").strip()
            if not text:
                missing_text_rows += 1
                continue
            valid_rows += 1

    if empty_rows:
        warnings.append(f"{empty_rows} linhas vazias foram ignoradas.")
    if missing_text_rows:
        warnings.append(f"{missing_text_rows} linhas sem texto foram ignoradas.")

    return CsvInspection(
        total_rows=total_rows,
        valid_rows=valid_rows,
        empty_rows=empty_rows,
        missing_text_rows=missing_text_rows,
        fieldnames=fieldnames,
        text_column=text_column or "",
        id_column=id_column or "",
        target_column=target_column or "",
        intent_column=intent_column or "",
        delimiter=getattr(dialect, "delimiter", ","),
        file_format="csv",
        sheet_name="",
        warnings=tuple(warnings),
    )


def iter_feedback(path: Path, limit: int | None = None) -> Iterable[CsvFeedback]:
    if _is_excel_path(path):
        yield from _iter_xlsx_feedback(path, limit=limit)
        return

    yielded = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        reader = csv.DictReader(handle, dialect=_sniff_dialect(sample))
        if not reader.fieldnames:
            return

        normalized = {_normalize_column(name): name for name in reader.fieldnames}
        text_column = _first_existing(normalized, TEXT_COLUMNS)
        if text_column is None:
            available = ", ".join(reader.fieldnames)
            raise ValueError(
                "CSV sem coluna de texto reconhecida. Use uma coluna chamada text, feedback, comment, "
                f"content, review, description ou message. Colunas encontradas: {available}"
            )

        id_column = _first_existing(normalized, ID_COLUMNS)
        target_column = _first_existing(normalized, TARGET_COLUMNS)
        intent_column = _first_existing(normalized, INTENT_COLUMNS)

        for index, row in enumerate(reader, start=1):
            if limit and yielded >= limit:
                return
            if not any(row.values()):
                continue
            text = (row.get(text_column) or "").strip()
            if not text:
                continue
            yielded += 1
            yield CsvFeedback(
                source_id=(row.get(id_column) or str(index)).strip() if id_column else str(index),
                text=text,
                target=(row.get(target_column) or "").strip() if target_column else "",
                intent=(row.get(intent_column) or "").strip() if intent_column else "",
            )


def _inspect_xlsx(path: Path, limit: int | None = None) -> CsvInspection:
    total_rows = 0
    valid_rows = 0
    empty_rows = 0
    missing_text_rows = 0
    warnings = []

    workbook = _load_xlsx_workbook(path)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        schema = _xlsx_schema(rows)
        fieldnames = schema.fieldnames
        text_column = schema.text_column
        id_column = schema.id_column
        target_column = schema.target_column
        intent_column = schema.intent_column
        warnings.extend(schema.warnings)

        if not id_column:
            warnings.append("Nenhuma coluna de ID reconhecida; o numero da linha sera usado como identificador.")
        if not target_column:
            warnings.append("Nenhuma coluna de alvo tecnico reconhecida; o alvo sera inferido pelo pipeline.")
        if not intent_column:
            warnings.append("Nenhuma coluna de intencao reconhecida; a intencao sera inferida pelo pipeline.")

        for row in _chain_rows(schema.initial_rows, rows):
            if limit and valid_rows >= limit:
                break
            values = _row_to_dict(fieldnames, row)
            if not any(values.values()):
                empty_rows += 1
                continue
            total_rows += 1
            text = values.get(text_column, "").strip()
            if not text:
                missing_text_rows += 1
                continue
            valid_rows += 1

        if empty_rows:
            warnings.append(f"{empty_rows} linhas vazias foram ignoradas.")
        if missing_text_rows:
            warnings.append(f"{missing_text_rows} linhas sem texto foram ignoradas.")

        return CsvInspection(
            total_rows=total_rows,
            valid_rows=valid_rows,
            empty_rows=empty_rows,
            missing_text_rows=missing_text_rows,
            fieldnames=fieldnames,
            text_column=text_column or "",
            id_column=id_column or "",
            target_column=target_column or "",
            intent_column=intent_column or "",
            delimiter="planilha",
            file_format="xlsx",
            sheet_name=worksheet.title,
            warnings=tuple(warnings),
        )
    finally:
        workbook.close()


def _iter_xlsx_feedback(path: Path, limit: int | None = None) -> Iterable[CsvFeedback]:
    yielded = 0
    workbook = _load_xlsx_workbook(path)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        schema = _xlsx_schema(rows)
        fieldnames = schema.fieldnames
        text_column = schema.text_column
        id_column = schema.id_column
        target_column = schema.target_column
        intent_column = schema.intent_column

        for index, row in enumerate(_chain_rows(schema.initial_rows, rows), start=1 if schema.initial_rows else 2):
            if limit and yielded >= limit:
                return
            values = _row_to_dict(fieldnames, row)
            if not any(values.values()):
                continue
            text = values.get(text_column, "").strip()
            if not text:
                continue
            yielded += 1
            yield CsvFeedback(
                source_id=values.get(id_column, "").strip() if id_column and values.get(id_column) else str(index),
                text=text,
                target=values.get(target_column, "").strip() if target_column else "",
                intent=values.get(intent_column, "").strip() if intent_column else "",
            )
    finally:
        workbook.close()


def _sniff_dialect(sample: str):
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        return csv.excel


def _normalize_column(name: str) -> str:
    return (name or "").lower().strip().replace("-", "_").replace(" ", "_")


def _first_existing(columns: dict[str, str], candidates: tuple[str, ...]) -> str | None:
    for candidate in candidates:
        if candidate in columns:
            return columns[candidate]
    return None


def _is_excel_path(path: Path) -> bool:
    return path.suffix.lower() in {".xlsx", ".xlsm"}


def _load_xlsx_workbook(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("A biblioteca openpyxl e necessaria para processar arquivos XLSX.") from exc
    return load_workbook(filename=path, read_only=True, data_only=True)


def _first_non_empty_row(rows):
    for row in rows:
        values = tuple(_cell_to_text(value) for value in row)
        if any(values):
            return values
    raise ValueError("Planilha vazia. Inclua uma linha de cabecalho e ao menos um feedback.")


def _xlsx_schema(rows) -> _XlsxSchema:
    first_row = _first_non_empty_row(rows)
    fieldnames = tuple(_cell_to_text(value) or f"column_{index}" for index, value in enumerate(first_row, start=1))
    normalized = {_normalize_column(name): name for name in fieldnames}
    text_column = _first_existing(normalized, TEXT_COLUMNS)
    id_column = _first_existing(normalized, ID_COLUMNS) or ""
    target_column = _first_existing(normalized, TARGET_COLUMNS) or ""
    intent_column = _first_existing(normalized, INTENT_COLUMNS) or ""

    if text_column is not None:
        return _XlsxSchema(
            fieldnames=fieldnames,
            text_column=text_column,
            id_column=id_column,
            target_column=target_column,
            intent_column=intent_column,
            initial_rows=(),
        )

    preview_rows = [first_row]
    for row in rows:
        values = tuple(_cell_to_text(value) for value in row)
        if any(values):
            preview_rows.append(values)
        if len(preview_rows) >= 30:
            break

    inferred_index = _infer_text_column_index(preview_rows)
    if inferred_index is None:
        available = ", ".join(fieldnames)
        raise ValueError(
            "Planilha sem coluna de texto reconhecida. Use uma coluna chamada text, feedback, comment, "
            f"content, review, description ou message. Colunas encontradas: {available}"
        )

    column_count = max(len(row) for row in preview_rows)
    inferred_fieldnames = tuple(f"column_{index}" for index in range(1, column_count + 1))
    text_column = inferred_fieldnames[inferred_index]
    return _XlsxSchema(
        fieldnames=inferred_fieldnames,
        text_column=text_column,
        id_column="",
        target_column="",
        intent_column="",
        initial_rows=tuple(preview_rows),
        warnings=(
            "Nenhum cabecalho reconhecido na planilha; "
            f"a coluna {inferred_index + 1} foi usada como texto do feedback.",
        ),
    )


def _infer_text_column_index(rows: list[tuple[str, ...]]) -> int | None:
    if not rows:
        return None

    column_count = max(len(row) for row in rows)
    best_index = None
    best_score = 0
    for index in range(column_count):
        score = sum(_text_cell_score(row[index] if index < len(row) else "") for row in rows)
        if score > best_score:
            best_score = score
            best_index = index

    return best_index if best_score > 0 else None


def _text_cell_score(value: str) -> int:
    text = _cell_to_text(value)
    if not text:
        return 0
    if _looks_numeric(text):
        return 0

    score = min(len(text), 240)
    if " " in text:
        score += 20
    if len(text) >= 25:
        score += 40
    return score


def _looks_numeric(text: str) -> bool:
    normalized = text.strip().replace(".", "", 1).replace(",", "", 1)
    return normalized.isdigit()


def _chain_rows(initial_rows: tuple[tuple[str, ...], ...], rows):
    yield from initial_rows
    for row in rows:
        yield row


def _row_to_dict(fieldnames: tuple[str, ...], row) -> dict[str, str]:
    values = [_cell_to_text(value) for value in row]
    if len(values) < len(fieldnames):
        values.extend([""] * (len(fieldnames) - len(values)))
    return {fieldname: values[index] if index < len(values) else "" for index, fieldname in enumerate(fieldnames)}


def _cell_to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
